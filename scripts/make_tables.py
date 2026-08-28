#!/usr/bin/env python3
"""Render publication tables and Supplementary Data from frozen artifacts."""

from __future__ import annotations

import json
import re
from datetime import UTC, datetime
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
TABLES = ROOT / "paper/tables"
SI_TABLES = ROOT / "paper/supplementary/tables"
SUPP_DATA = ROOT / "paper/supplementary_data"
ED = ROOT / "paper/extended_data"


def latex_escape(value: object) -> str:
    text = str(value)
    for source, target in (
        ("&", r"\&"),
        ("%", r"\%"),
        ("_", r"\_"),
        ("#", r"\#"),
    ):
        text = text.replace(source, target)
    return text


def latex_table(frame: pd.DataFrame, path: Path, columns: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [rf"\begin{{tabular}}{{{columns}}}", r"\toprule"]
    lines.append(" & ".join(latex_escape(c) for c in frame.columns) + r" \\")
    lines.append(r"\midrule")
    for row in frame.itertuples(index=False, name=None):
        lines.append(" & ".join(latex_escape(value) for value in row) + r" \\")
    lines.extend([r"\bottomrule", r"\end{tabular}"])
    path.write_text("\n".join(lines) + "\n")


def workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        frozen_time = datetime(2026, 8, 28, tzinfo=UTC)
        writer.book.properties.created = frozen_time
        writer.book.properties.modified = frozen_time
        for name, frame in sheets.items():
            sheet = name[:31]
            frame.to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.book[sheet]
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1769AA")
                cell.alignment = Alignment(wrap_text=True)
            ws.freeze_panes = "A2"
            for column in ws.columns:
                values = [str(cell.value or "") for cell in column[:1000]]
                ws.column_dimensions[column[0].column_letter].width = min(
                    58, max(10, max(map(len, values)) + 2)
                )
    normalized = path.with_name(f".{path.name}.normalized")
    with (
        ZipFile(path) as source,
        ZipFile(normalized, "w", compression=ZIP_DEFLATED, compresslevel=9) as target,
    ):
        for name in sorted(source.namelist()):
            payload = source.read(name)
            if name == "docProps/core.xml":
                text = payload.decode("utf-8")
                text = re.sub(
                    r"(<dcterms:(?:created|modified)[^>]*>).*?(</dcterms:(?:created|modified)>)",
                    r"\g<1>2026-08-28T00:00:00Z\g<2>",
                    text,
                )
                payload = text.encode("utf-8")
            original = source.getinfo(name)
            info = ZipInfo(name, date_time=(2026, 8, 28, 0, 0, 0))
            info.compress_type = ZIP_DEFLATED
            info.external_attr = original.external_attr
            info.internal_attr = original.internal_attr
            info.create_system = original.create_system
            target.writestr(info, payload, compress_type=ZIP_DEFLATED, compresslevel=9)
    normalized.replace(path)


def response_priors() -> pd.DataFrame:
    rows = [
        (
            1,
            "combisolv_qm",
            "COSMOtherm water response",
            "CombiSolv-QM",
            "kcal mol-1",
            "CHEMELEON D-MPNN",
            3959,
            "direct",
        ),
        (
            2,
            "abraham_e",
            "excess molar refraction E",
            "SoluteML",
            "Abraham scale",
            "ExtraTrees",
            8098,
            "direct",
        ),
        (
            3,
            "abraham_s",
            "dipolarity/polarizability S",
            "SoluteML",
            "Abraham scale",
            "ExtraTrees",
            8098,
            "direct",
        ),
        (
            4,
            "abraham_a",
            "hydrogen-bond acidity A",
            "SoluteML",
            "Abraham scale",
            "ExtraTrees",
            8098,
            "direct",
        ),
        (
            5,
            "abraham_b",
            "hydrogen-bond basicity B",
            "SoluteML",
            "Abraham scale",
            "ExtraTrees",
            8098,
            "direct",
        ),
        (
            6,
            "abraham_l",
            "hexadecane-air partition L",
            "SoluteML",
            "Abraham scale",
            "ExtraTrees",
            8098,
            "direct",
        ),
        (
            7,
            "openff_corrected",
            "explicit-water alchemical response",
            "OpenFF 2.3.0 ASFE",
            "kcal mol-1",
            "ExtraTrees",
            520,
            "prediction + residual",
        ),
        (
            8,
            "gbn2_corrected",
            "implicit-solvent hydration response",
            "GBn2",
            "kcal mol-1",
            "ExtraTrees",
            550,
            "prediction + residual",
        ),
        (
            9,
            "smd_water",
            "SMD water response",
            "MolSolv",
            "kcal mol-1",
            "CHEMELEON D-MPNN",
            350359,
            "direct",
        ),
        (
            10,
            "conf_gas_corr",
            "gas conformer correction",
            "ConfSolv H2O",
            "kcal mol-1",
            "LightGBM",
            17829,
            "direct",
        ),
        (
            11,
            "conf_solution_corr",
            "solution conformer correction",
            "ConfSolv H2O",
            "kcal mol-1",
            "LightGBM",
            17829,
            "direct",
        ),
        (
            12,
            "conf_hydration_corr",
            "hydration conformer correction",
            "ConfSolv H2O",
            "kcal mol-1",
            "LightGBM",
            17829,
            "direct",
        ),
        (
            13,
            "conf_gsolv_sd",
            "conformer solvation-energy spread",
            "ConfSolv H2O",
            "kcal mol-1",
            "LightGBM",
            17829,
            "direct",
        ),
        (
            14,
            "conf_response_mean",
            "mean conformer solvent response",
            "ConfSolv H2O",
            "kcal mol-1",
            "LightGBM",
            17829,
            "direct",
        ),
        (
            15,
            "conf_response_sd",
            "conformer response spread",
            "ConfSolv H2O",
            "kcal mol-1",
            "LightGBM",
            17829,
            "direct",
        ),
    ]
    columns = [
        "prior",
        "name",
        "physical meaning",
        "source",
        "units",
        "surrogate",
        "training rows",
        "transformation",
    ]
    frame = pd.DataFrame(rows, columns=columns)
    frame["inference"] = "structure surrogate; no simulation"
    return frame


def endpoint_sources() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("FreeSolv only", 69, "benchmark-disjoint connectivity"),
            ("CombiSolv-Exp only", 588, "water records; benchmark-disjoint connectivity"),
            ("FreeSolv + CombiSolv-Exp", 490, "one duplicate-resolved connectivity"),
            ("SoluteML dGsolvDB3", 133, "at least two source measurements"),
        ],
        columns=["source group", "selected records", "selection rule"],
    )


def experiment_ledger(metrics: dict) -> pd.DataFrame:
    m = metrics["methods"]
    rows = [
        (
            "Matched structure-only",
            "Do response priors add information?",
            "none",
            "A",
            "fixed five-fold OOF",
            m["matched_structure_only"]["mae_kcal_mol"],
            "matched control",
            "results/confirmatory/standardized_exclusion_endpoint_predictions.parquet",
        ),
        (
            "Empirical/residual block",
            "Do compact empirical axes and corrected responses help?",
            "Abraham, OpenFF, GBn2",
            "B",
            "fixed five-fold OOF",
            m["empirical_residual_block"]["mae_kcal_mol"],
            "component",
            "results/confirmatory/standardized_exclusion_endpoint_predictions.parquet",
        ),
        (
            "Computation core",
            "Do raw computed responses help?",
            "CombiSolv-QM, OpenFF, GBn2",
            "C",
            "fixed five-fold OOF",
            m["computation_core_block"]["mae_kcal_mol"],
            "component",
            "results/confirmatory/standardized_exclusion_endpoint_predictions.parquet",
        ),
        (
            "SMD water",
            "Does water-specific continuum response transfer?",
            "MolSolv SMD",
            "D",
            "fixed five-fold OOF",
            m["smd_water_block"]["mae_kcal_mol"],
            "component",
            "results/confirmatory/standardized_exclusion_endpoint_predictions.parquet",
        ),
        (
            "ConfSolv",
            "Do conformational response summaries transfer?",
            "ConfSolv H2O",
            "E",
            "fixed five-fold OOF",
            m["confsolv_block"]["mae_kcal_mol"],
            "component",
            "results/confirmatory/standardized_exclusion_endpoint_predictions.parquet",
        ),
        (
            "Full SolvAI",
            "Do aligned response priors improve the matched endpoint?",
            "15 priors",
            "F",
            "fixed five-fold OOF",
            m["full_solvai"]["mae_kcal_mol"],
            "final model",
            "results/confirmatory/standardized_exclusion_endpoint_predictions.parquet",
        ),
        (
            "Shuffled priors",
            "Does molecule-response alignment matter?",
            "15 permuted priors",
            "negative control",
            "five permutations × six partitions",
            None,
            "evaluation only",
            "results/confirmatory/standardized_exclusion_endpoint_shuffle_predictions.parquet",
        ),
        (
            "Zero-ARROW transfer",
            "Do priors help without ARROW endpoint labels?",
            "15 priors",
            "transfer control",
            "all 85 as test",
            m["zero_arrow_full_solvai"]["mae_kcal_mol"],
            "evaluation only",
            "results/confirmatory/standardized_exclusion_endpoint_predictions.parquet",
        ),
    ]
    for regime, values in metrics["global_separation"].items():
        rows.append(
            (
                regime,
                "Does the prior advantage survive global chemical separation?",
                "15 priors",
                "separation control",
                regime,
                values["F_full_solvai"]["mae_kcal_mol"],
                "evaluation only",
                "results/confirmatory/standardized_exclusion_global_separation_predictions.parquet",
            )
        )
    for name, value in metrics["alternative_supervision"].items():
        rows.append(
            (
                name.replace("_", " "),
                "Can an alternative physical representation improve transfer?",
                name,
                "exploratory",
                "campaign OOF screen",
                value,
                "not retained",
                "results/ablations/",
            )
        )
    for name, value in metrics["multilambda"]["method_mae_kcal_mol"].items():
        rows.append(
            (
                name,
                "Can predicted lambda response improve the endpoint?",
                "PIMD2 response",
                "exploratory",
                "matched OOF",
                value,
                "not retained",
                "results/ablations/multilambda_metrics.csv",
            )
        )
    return pd.DataFrame(
        rows,
        columns=[
            "experiment",
            "hypothesis",
            "physical source",
            "analysis ID",
            "evaluation",
            "MAE (kcal/mol)",
            "role in final model",
            "artifact",
        ],
    )


def main() -> None:
    metrics = json.loads((ROOT / "results/paper_metrics.json").read_text())
    methods = metrics["methods"]
    expected = {
        "matched_structure_only": 0.3033484655954973,
        "full_solvai": 0.20223406721910991,
        "arrow_pimd8": 0.20483647058823526,
    }
    for key, value in expected.items():
        if abs(methods[key]["mae_kcal_mol"] - value) > 1e-12:
            raise AssertionError(f"Refusing to render stale metric {key}")
    for folder in (TABLES, SI_TABLES, SUPP_DATA, ED):
        folder.mkdir(parents=True, exist_ok=True)

    repeats = metrics["repeated_splits"]
    global_results = metrics["global_separation"]
    paired = pd.DataFrame(metrics["paired_confirmatory"])
    primary_pair = paired.loc[paired.analysis.eq("primary_F_full_solvai")].iloc[0]
    shuffle_pair = paired.loc[paired.analysis.eq("aligned_vs_mean_shuffle_primary_-1")].iloc[0]
    macros = {
        "BenchmarkN": 85,
        "ClassicalMAE": f"{methods['classical_arrow']['mae_kcal_mol']:.3f}",
        "PIMDMAE": f"{methods['arrow_pimd8']['mae_kcal_mol']:.3f}",
        "MatchedMAE": f"{methods['matched_structure_only']['mae_kcal_mol']:.3f}",
        "SolvAIMAE": f"{methods['full_solvai']['mae_kcal_mol']:.3f}",
        "SolvAIDelta": f"{primary_pair.difference:.3f}",
        "SolvAICILow": f"{primary_pair.ci_low_95:.3f}",
        "SolvAICIHigh": f"{primary_pair.ci_high_95:.3f}",
        "RepeatMean": f"{repeats['full_solvai']['mean_kcal_mol']:.3f}",
        "RepeatSD": f"{repeats['full_solvai']['sd_kcal_mol']:.3f}",
        "MatchedRepeatMean": f"{repeats['matched_structure_only']['mean_kcal_mol']:.3f}",
        "NarrowMAE": f"{methods['narrow_response']['mae_kcal_mol']:.3f}",
        "NarrowSMDMAE": f"{methods['narrow_plus_smd']['mae_kcal_mol']:.3f}",
        "ShuffledMAE": f"{shuffle_pair.reference_mae:.3f}",
        "FamilyMAE": f"{global_results['global_family']['F_full_solvai']['mae_kcal_mol']:.3f}",
        "ScaffoldMAE": f"{global_results['global_scaffold']['F_full_solvai']['mae_kcal_mol']:.3f}",
        "ClusterMAE": f"{global_results['global_butina_0_70']['F_full_solvai']['mae_kcal_mol']:.3f}",
        "NNSeventyMatchedMAE": f"{global_results['global_nn_0.70']['A_structure_only']['mae_kcal_mol']:.3f}",
        "NNSeventySolvAIMAE": f"{global_results['global_nn_0.70']['F_full_solvai']['mae_kcal_mol']:.3f}",
        "FamilyMatchedMAE": f"{global_results['global_family']['A_structure_only']['mae_kcal_mol']:.3f}",
        "ScaffoldMatchedMAE": f"{global_results['global_scaffold']['A_structure_only']['mae_kcal_mol']:.3f}",
        "ZeroArrowMatchedMAE": f"{methods['zero_arrow_structure_only']['mae_kcal_mol']:.3f}",
        "ZeroArrowMAE": f"{methods['zero_arrow_full_solvai']['mae_kcal_mol']:.3f}",
        "MolSolvN": f"{metrics['data_counts']['molsolv_confirmatory_rows']:,}",
        "ConfSolvN": f"{metrics['data_counts']['confsolv_confirmatory_rows']:,}",
    }
    (TABLES / "metrics_macros.tex").write_text(
        "".join(f"\\newcommand{{\\{key}}}{{{value}}}\n" for key, value in macros.items())
    )

    comparison = pd.DataFrame(
        [
            ("Classical ARROW", "yes", methods["classical_arrow"]["mae_kcal_mol"], "—", "—"),
            ("ARROW/PIMD8", "yes", methods["arrow_pimd8"]["mae_kcal_mol"], "—", "—"),
            (
                "Matched structure-only",
                "no",
                methods["matched_structure_only"]["mae_kcal_mol"],
                "—",
                "—",
            ),
            (
                "SolvAI",
                "no",
                methods["full_solvai"]["mae_kcal_mol"],
                global_results["global_family"]["F_full_solvai"]["mae_kcal_mol"],
                global_results["global_scaffold"]["F_full_solvai"]["mae_kcal_mol"],
            ),
            (
                "SolvAI, zero ARROW labels",
                "no",
                methods["zero_arrow_full_solvai"]["mae_kcal_mol"],
                "—",
                "—",
            ),
        ],
        columns=[
            "Method",
            "Simulation at inference",
            "Fixed OOF MAE",
            "Global family MAE",
            "Global scaffold MAE",
        ],
    )
    comparison.to_csv(TABLES / "main_comparison.csv", index=False)
    comparison.to_csv(ROOT / "results/model_comparison.csv", index=False)
    display = comparison.copy()
    for column in display.columns[2:]:
        display[column] = display[column].map(
            lambda value: f"{value:.3f}" if isinstance(value, float) else value
        )
    latex_table(display, TABLES / "main_comparison.tex", "lcccc")
    comparison.to_csv(ED / "ED_Table1.csv", index=False)
    latex_table(display, ED / "ED_Table1.tex", "lcccc")

    priors = response_priors()
    endpoint = endpoint_sources()
    source_summary = pd.DataFrame(
        [
            ("CombiSolv-QM", "unique structures", 3961, 2, 3959, "COSMOtherm water"),
            ("MolSolv", "SMD calculations", 350391, 32, 350359, "SMD(water)"),
            ("ConfSolv", "model-usable connectivities", 17851, 22, 17829, "conformer response"),
            (
                "Endpoint labels",
                "experimental connectivities",
                1280,
                0,
                1280,
                "hydration free energy",
            ),
        ],
        columns=["source", "unit", "before standardized exclusion", "removed", "retained", "role"],
    )
    latex_table(priors, SI_TABLES / "response_priors.tex", "rlllllrrl")
    latex_table(source_summary, SI_TABLES / "source_provenance.tex", "llrrrp{3.0cm}")
    latex_table(endpoint, SI_TABLES / "endpoint_sources.tex", "lrl")

    repeat_rows = []
    repeat_metrics = pd.read_csv(
        ROOT / "results/confirmatory/standardized_exclusion_endpoint_metrics.csv"
    )
    for row in repeat_metrics.loc[
        repeat_metrics.partition.eq("standardized_exclusion_repeat")
    ].itertuples():
        repeat_rows.append((row.repeat, int(row.split_seed), row.method, f"{row.mae:.5f}"))
    repeat_table = pd.DataFrame(repeat_rows, columns=["Repeat", "Split seed", "Method", "MAE"])
    latex_table(repeat_table, SI_TABLES / "repeat_values.tex", "rrlr")

    separation_table = pd.read_csv(
        ROOT / "results/confirmatory/standardized_exclusion_global_separation_metrics.csv"
    )
    separation_table["mae"] = separation_table.mae.map(lambda value: f"{value:.3f}")
    latex_table(
        separation_table[["regime", "method", "n", "mae"]],
        SI_TABLES / "global_separation.tex",
        "llrr",
    )

    family = pd.DataFrame(metrics["chemistry_family"])
    family["mae_kcal_mol"] = family.mae_kcal_mol.map(lambda value: f"{value:.3f}")
    latex_table(family, SI_TABLES / "family_errors.tex", "lrr")
    manifest = json.loads((ROOT / "models/final/manifest.json").read_text())["model_files"]
    artifacts = pd.DataFrame([{"file": key, **value} for key, value in manifest.items()])
    latex_table(artifacts, SI_TABLES / "artifact_manifest.tex", "lrl")

    runtime = json.loads((ROOT / "results/runtime/runtime_benchmark.json").read_text())
    runtime_rows = pd.DataFrame(
        [
            ("cold single molecule", f"{runtime['single_molecule']['cold_seconds']:.3f} s"),
            (
                "warm single molecule median",
                f"{runtime['single_molecule']['warm_median_seconds']:.3f} s",
            ),
            ("warm single molecule p95", f"{runtime['single_molecule']['warm_p95_seconds']:.3f} s"),
            ("batch 32 median", f"{runtime['batch']['median_seconds']:.3f} s"),
            ("batch 32 per molecule", f"{runtime['batch']['median_seconds_per_molecule']:.3f} s"),
        ],
        columns=["Measurement", "Value"],
    )
    latex_table(runtime_rows, SI_TABLES / "runtime.tex", "ll")

    ledger = experiment_ledger(metrics)
    ledger.to_csv(SUPP_DATA / "Supplementary_Data_1_experiment_ledger.csv", index=False)
    workbook(SUPP_DATA / "Supplementary_Data_1_experiment_ledger.xlsx", {"experiments": ledger})

    predictions = pd.read_parquet(
        ROOT / "results/confirmatory/standardized_exclusion_endpoint_predictions.parquet"
    )
    primary = predictions.loc[predictions.partition.eq("standardized_exclusion_primary")].copy()
    base = primary[
        [
            "molecule_id",
            "molecule_name",
            "canonical_smiles",
            "functional_group_family",
            "scaffold",
            "fold",
            "y_true",
        ]
    ].drop_duplicates("molecule_id")
    benchmark = pd.read_parquet(ROOT / "data/benchmark/arrow_solvation_master.parquet")
    benchmark = benchmark.loc[
        benchmark.solvent.eq("water"), ["molecule_id", "delta_g_pimd8", "delta_g_classical_arrow"]
    ]
    wide = primary.pivot(index="molecule_id", columns="method", values="y_pred").reset_index()
    molecule_predictions = base.merge(benchmark, on="molecule_id", validate="one_to_one").merge(
        wide, on="molecule_id", validate="one_to_one"
    )
    molecule_predictions.to_csv(
        SUPP_DATA / "Supplementary_Data_2_molecule_predictions.csv", index=False
    )
    workbook(
        SUPP_DATA / "Supplementary_Data_2_molecule_predictions.xlsx",
        {"primary predictions": molecule_predictions},
    )

    assignments = pd.read_parquet(
        ROOT / "results/confirmatory/standardized_exclusion_chemical_separation_assignments.parquet"
    )
    repeat_assignments = predictions.loc[
        predictions.partition.eq("standardized_exclusion_repeat")
        & predictions.method.eq("F_full_solvai"),
        ["molecule_id", "repeat", "split_seed", "fold"],
    ]
    assignments.to_csv(SUPP_DATA / "Supplementary_Data_3_split_assignments.csv", index=False)
    workbook(
        SUPP_DATA / "Supplementary_Data_3_split_assignments.xlsx",
        {"chemical assignments": assignments, "repeat folds": repeat_assignments},
    )

    audit_summary = pd.read_csv(ROOT / "audits/confirmatory/chemical_distance_summary.csv")
    source_summary.to_csv(SUPP_DATA / "Supplementary_Data_4_teacher_sources.csv", index=False)
    priors.to_csv(SUPP_DATA / "Supplementary_Data_4_teacher_priors.csv", index=False)
    endpoint.to_csv(SUPP_DATA / "Supplementary_Data_4_endpoint_sources.csv", index=False)
    workbook(
        SUPP_DATA / "Supplementary_Data_4_teacher_manifests.xlsx",
        {
            "response priors": priors,
            "teacher sources": source_summary,
            "endpoint sources": endpoint,
            "identity audit": audit_summary,
        },
    )
    print("Rendered confirmatory manuscript tables and Supplementary Data.")


if __name__ == "__main__":
    main()
